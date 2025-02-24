import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from typing import List, Any
from pydantic import BaseModel
import mysql.connector
import os
from dotenv import load_dotenv

# Carregar variáveis do .env
load_dotenv()

class ProductSchema(BaseModel):
    id: int
    title: str
    price: float
    category: str
    description: str
    image: str
    timestamp: Any = None
    rating: Any = None


# Configuração da conexão usando variáveis do .env
def get_db_connection():
    return mysql.connector.connect(
        host=os.getenv("DB_HOST"),
        user=os.getenv("DB_USER"),
        password=os.getenv("DB_PASSWORD"),
        database=os.getenv("DB_NAME"),
        auth_plugin=os.getenv("DB_AUTH_PLUGIN")
    )

# Busca produtos no banco de dados
def buscar_produtos() -> List[ProductSchema]:
    conexao = get_db_connection()
    cursor = conexao.cursor()
    query = "SELECT id, title, price, category, description, image FROM produtos;"
    cursor.execute(query)
    resultado = cursor.fetchall()
    produtos = [ProductSchema(id=linha[0],
                          title=linha[1],
                          price=linha[2],
                          category=linha[3],
                          description=linha[4],
                          image=linha[5]) for linha in resultado]

    cursor.close()
    conexao.close()
    return produtos

# Gera o arquivo Excel formatado
def gerar_excel(produtos: List[ProductSchema], nome_arquivo: str = "relatorio_produtos.xlsx"):
    if not produtos:
        print("Nenhum produto encontrado.")
        return

    df = pd.DataFrame([p.model_dump() for p in produtos])
    df.to_excel(nome_arquivo, index=False, sheet_name="Produtos")

    wb = load_workbook(nome_arquivo)
    ws = wb["Produtos"]

    fill_vermelho = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value > 100:
                cell.fill = fill_vermelho

    ws_stats = wb.create_sheet("Estatísticas")
    estatisticas = df.groupby("category")["price"].mean().reset_index()
    estatisticas.columns = ["Categoria", "Preço Médio"]

    ws_stats.append(["Categoria", "Preço Médio"])
    for _, row in estatisticas.iterrows():
        ws_stats.append(row.tolist())

    for sheet in [ws, ws_stats]:
        for col in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    wb.save(nome_arquivo)
    print(f"✅ Arquivo '{nome_arquivo}' gerado com sucesso!")

if __name__ == "__main__":
    produtos = buscar_produtos()
    gerar_excel(produtos)
