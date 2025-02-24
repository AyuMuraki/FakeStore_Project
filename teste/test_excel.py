import pytest
from openpyxl import load_workbook

def test_excel_gerado():
    nome_arquivo = "relatorio_produtos.xlsx"

    # Carregar o arquivo Excel gerado
    wb = load_workbook(nome_arquivo)
    ws = wb["Produtos"]

    # Verificar se as colunas estão corretas
    colunas_esperadas = ["id", "title", "price", "category", "description", "image", "timestamp"]
    colunas = [cell.value for cell in ws[1]]  # Primeira linha (cabeçalho)
    assert colunas == colunas_esperadas

    # Verificar a existência da aba "Estatísticas"
    assert "Estatísticas" in wb.sheetnames
    ws_stats = wb["Estatísticas"]
    
    # Verificar se há estatísticas de preço médio por categoria
    assert ws_stats.cell(row=2, column=1).value == "jewelery"  # Exemplo de categoria
    assert isinstance(ws_stats.cell(row=2, column=2).value, (int, float))  # Preço médio
    
    wb.close()
